#!/usr/bin/env python3
"""
generate_dashboard.py
Reads the Push Films accounts spreadsheet and writes dashboard_data.json.

Run after any spreadsheet update, then:
    rm -f ~/push-films-website/.git/HEAD.lock && cd ~/push-films-website && \
    git add dashboard_data.json && git commit -m "Update dashboard data" && git push
"""

import os, glob, json, re
from datetime import datetime
import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────────────
SPREADSHEET_DIR = os.path.expanduser(
    '~/Library/Mobile Documents/com~apple~CloudDocs/'
    'ICLOUD/GREG PERSONAL /PERSONAL FINANCES /CLAUDE VERSION - MARCH 26 /'
)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_OUT    = os.path.join(SCRIPT_DIR, 'dashboard_data.json')

# ── Find most recently modified ACCOUNTS*.xlsx ────────────────────────────────
xlsx_files = glob.glob(os.path.join(SPREADSHEET_DIR, 'ACCOUNTS*.xlsx'))
if not xlsx_files:
    raise FileNotFoundError(f"No ACCOUNTS*.xlsx found in:\n  {SPREADSHEET_DIR}")
spreadsheet_path = max(xlsx_files, key=os.path.getmtime)
print(f"Reading: {os.path.basename(spreadsheet_path)}")

# ── Load existing JSON — preserve manually annotated fields ───────────────────
existing_atl = {}
if os.path.exists(JSON_OUT):
    with open(JSON_OUT, 'r') as f:
        old = json.load(f)
    for entry in old.get('atl', []):
        existing_atl[entry['m']] = entry

# ── Load workbook (data_only reads cached formula values) ─────────────────────
wb = openpyxl.load_workbook(spreadsheet_path, data_only=True)

# ── Helpers ───────────────────────────────────────────────────────────────────
def norm(t):
    return re.sub(r'\s+', ' ', str(t or '').lower()).strip()

def find_col(ws, row, *keywords):
    """Scan a header row for the first column containing any keyword."""
    for keyword in keywords:
        for col in range(1, min(ws.max_column + 1, 25)):
            if keyword in norm(ws.cell(row=row, column=col).value or ''):
                return col
    return None

# ── Payee name lookup — maps raw bank memo fragments to clean display names ───
PAYEE_MAP = [
    # Pattern (upper)           Clean name
    ('IMAGINATION',             'Imagination Europe'),
    ('WORKSOME',                'Imagination Europe'),
    ('CHARTWELL',               'Chartwell Travel'),
    ('BRIGHTLIGHT',             'Brightlight Films'),
    ('TALLER STORIES',          'Taller Stories'),
    ('EYE RISE',                'Eye Rise Films'),
    ('8 FEET',                  '8 Feet Fixer'),
    ('BURNS',                   'Burns + Wilcox'),
    ('COASTHOUSE',              'Coasthouse Productions'),
    ('LETTICE',                 'Lettice Yardley'),
    ('PAUL BARTON',             'Paul Barton Cinema'),
    ('DAVID BRILL',             'David Brill'),
    ('BRAZN',                   'BRAZN'),
    ('DRUCE',                   'Charlie Druce'),
    ('CKM',                     'Charlie Druce'),
    ('ALASTAIR FOX',            'Alastair Fox'),
    ('WISE',                    'Alastair Fox'),
    ('SALLY HOBDEN',            'Sally Hobden'),
    ('CHESTERFIELD',            'Chesterfield Group'),
    ('PROTAPE',                 'Protape'),
    ('DYNAMIC DOX',             'Dynamic Dox'),
    ('NO ORDINARY',             'No Ordinary Films'),
    ('NOF',                     'No Ordinary Films'),
    ('NATURE OF THE',           'Nature of the Beast'),
    ('8 FEET PART',             '8 Feet Fixer'),
    ('VITESSE',                 'Vitesse'),
    ('THEOPENLAB',              'TheOpenLab'),
    ('OPENLAB',                 'TheOpenLab'),
    # CRH / international wire patterns
    ('LAME PRODU',              'LAME'),
    ('LAME CR',                 'LAME'),
    ('LAME FINAL',              'LAME'),
    ('CRH IMAG',                'Imagination Europe'),
    ('CRH USA',                 'Imagination Europe'),
    ('PHOENIXC',                'Imagination Europe'),
    ('DUBLINCRH',               'Imagination Europe'),
    ('TALLER STO',              'Taller Stories'),
    ('HARTSHORN',               'DL Hartshorn'),
]

PAYMENT_TYPES = {'direct debit', 'standing order', 'funds transfer',
                 'card purchase', 'faster payment', 'bacs', 'bgc', 'crh',
                 'debit', 'credit', 'transfer', 'counter credit',
                 'unpaid direct debit', 'faster payments',
                 'contactless card purchase', 'charges commission for period'}

def extract_memo(ws, row):
    """Extract the best memo string from a data row (cols 4–7)."""
    candidates = []
    for c in range(4, 8):
        v = ws.cell(row=row, column=c).value
        if not v or not isinstance(v, str):
            continue
        cleaned = re.sub(r'\s+', ' ', v).strip()
        if not cleaned:
            continue
        if norm(cleaned) in PAYMENT_TYPES:
            continue
        candidates.append(cleaned)
    return candidates[0] if candidates else ''

def clean_payee(memo):
    """Map a raw bank memo to a readable payee name."""
    upper = memo.upper()
    for pattern, name in PAYEE_MAP:
        if pattern in upper:
            return name
    # Fall back to first 30 chars of raw memo, title-cased
    return memo[:40].strip().title() if memo else 'Unknown'

# ── Read income AND costs from each monthly tab ───────────────────────────────
# Tab names: "PUSH + 108 SEP 2024" etc.  Month label = "SEP 2024"
TAB_RE = re.compile(r'^PUSH \+ \d+ (.+)$')

income_by_month        = {}
costs_breakdown_by_month = {}  # list of {p, a} per month
payees_by_month        = {}   # dict of {payee_name: total} per month

_headers_printed = False
for sheet_name in wb.sheetnames:
    m = TAB_RE.match(sheet_name)
    if not m:
        continue
    month_label = m.group(1).strip()
    ws = wb[sheet_name]

    # Print column headers from row 3 for the first monthly tab (diagnostic)
    if not _headers_printed:
        _headers_printed = True
        print(f"\nRow 3 headers in '{sheet_name}':")
        for col in range(1, min(ws.max_column + 1, 20)):
            v = ws.cell(row=3, column=col).value
            if v is not None:
                print(f"  col {col} ({chr(64+col)}): {repr(v)}")
        # Also print first data row
        print("  First data row (row 4):")
        for col in range(1, min(ws.max_column + 1, 20)):
            v = ws.cell(row=4, column=col).value
            if v is not None:
                print(f"    col {col} ({chr(64+col)}): {repr(v)}")

    # Read ALL column headers from row 3 — the headers ARE the payee/category names
    all_headers = {}  # col -> cleaned header string
    for col in range(1, min(ws.max_column + 1, 40)):
        v = ws.cell(row=3, column=col).value
        if v is not None:
            h = str(v).strip().replace('\n', ' ').strip()
            if h:
                all_headers[col] = h

    # Identify income column
    income_col = None
    for col, h in all_headers.items():
        if any(k in h.lower() for k in ('income', 'funds in', 'push income', 'turnover')):
            income_col = col
            break

    # Only the "Project Costs" column is used for the breakdown.
    PRODUCTION_COLS = {}  # col -> display name
    for col, h in all_headers.items():
        if col == income_col:
            continue
        if 'project cost' in h.lower() or 'crh cost' in h.lower():
            PRODUCTION_COLS[col] = 'Project Costs'
            break

    # Date is in col B (col 2) for each data row
    DATE_COL = 2

    income_total = 0.0
    cost_items   = []  # list of {p: "DD MMM", a: amount}

    for row in range(4, ws.max_row + 1):
        # Col C gate for income — prevents formula/subtotal rows being double-counted
        bank_cell = ws.cell(row=row, column=3).value
        if income_col and bank_cell not in (None, '', 0):
            val = ws.cell(row=row, column=income_col).value
            if val and isinstance(val, (int, float)) and val > 0:
                income_total += val

        # Project costs — gate on col B having any non-empty value (date or string)
        # Some months store dates as strings rather than Excel date objects
        date_val = ws.cell(row=row, column=DATE_COL).value
        if date_val is None or date_val == '':
            continue  # no date at all = skip (formula/total row)

        if hasattr(date_val, 'strftime'):
            date_label = date_val.strftime('%-d %b')  # Excel date object → "14 Oct"
        else:
            # String date — try to parse, fall back to raw string
            date_str = str(date_val).strip()
            if not date_str:
                continue
            parsed = False
            for fmt in ('%Y-%m-%d %H:%M:%S', '%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y'):
                try:
                    from datetime import datetime as _dt
                    date_label = _dt.strptime(date_str.split()[0], fmt.split()[0]).strftime('%-d %b')
                    parsed = True
                    break
                except ValueError:
                    continue
            if not parsed:
                date_label = date_str[:10]

        for col, payee in PRODUCTION_COLS.items():
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, (int, float)) and val > 0:
                cost_items.append({'p': date_label, 'a': round(float(val), 2)})
                memo        = extract_memo(ws, row)
                payee_name  = clean_payee(memo)
                payee_totals = payees_by_month.setdefault(month_label, {})
                payee_totals[payee_name] = round(
                    payee_totals.get(payee_name, 0.0) + float(val), 2
                )

    income_by_month[month_label] = round(income_total, 2)
    if cost_items:
        costs_breakdown_by_month[month_label] = cost_items

# ── Read SUMMARY sheet ────────────────────────────────────────────────────────
ws_sum = wb['SUMMARY']

# ── Monthly totals — rows 4-22 ────────────────────────────────────────────────
# Expected columns: A=Month B=Spend C=Running D=vs Average E=Food F=Accom
monthly = []
for r in range(4, 23):
    month = ws_sum.cell(row=r, column=1).value
    if not month or str(month).strip() in ('TOTAL', 'MONTHLY AVG', ''):
        continue
    monthly.append({
        'm':       str(month).strip(),
        'spend':   round(float(ws_sum.cell(row=r, column=2).value or 0), 2),
        'running': round(float(ws_sum.cell(row=r, column=3).value or 0), 2),
        'vs':      round(float(ws_sum.cell(row=r, column=4).value or 0), 2),
        'food':    round(float(ws_sum.cell(row=r, column=5).value or 0), 2),
        'accom':   round(float(ws_sum.cell(row=r, column=6).value or 0), 2),
    })

# ── Category breakdown — DYNAMIC column discovery ────────────────────────────
# Header expected in row 27; data in rows 28-46.
# Reads ALL named columns — picks up new categories (Mortgage/Rent, Ocado etc.)
# automatically whenever Greg adds them to the spreadsheet.
NON_CAT = {'', 'total', 'total spend', 'monthly avg', 'monthly average', 'avg', 'month', 'other / uncategorised', 'other/uncategorised', 'uncategorised', 'other'}
CAT_COLS = {}
for col in range(2, 25):
    v = ws_sum.cell(row=27, column=col).value
    if v and norm(v) not in NON_CAT:
        CAT_COLS[str(v).strip()] = col

if CAT_COLS:
    print(f"\nCategories found: {list(CAT_COLS.keys())}")
else:
    # Fallback to hardcoded columns if header row is blank
    CAT_COLS = {
        'Food & Drink': 3, 'Transport': 4, 'Comms': 5, 'Home': 6,
        'Finance/Work': 7, 'Lifestyle': 8, 'Personal': 9, 'Family': 10,
    }
    print("\nWarning: category header row 27 appears empty — using hardcoded columns")

cats = []
for r in range(28, 47):
    month = ws_sum.cell(row=r, column=1).value
    if not month or str(month).strip() in ('TOTAL', 'MONTHLY AVG', ''):
        continue
    entry = {'m': str(month).strip()}
    for cat, col in CAT_COLS.items():
        val = ws_sum.cell(row=r, column=col).value
        entry[cat] = round(float(val or 0), 2)
    cats.append(entry)

# ── ATL project costs total — rows 53-71: A=Month C=Project Costs ─────────────
atl_costs = {}
for r in range(53, 72):
    month = ws_sum.cell(row=r, column=1).value
    if not month or str(month).strip() in ('TOTAL', ''):
        continue
    costs = ws_sum.cell(row=r, column=3).value or 0
    atl_costs[str(month).strip()] = round(float(costs), 2)

# ── Build ATL array ───────────────────────────────────────────────────────────
# income          → fresh from monthly tabs
# costs           → from SUMMARY col C
# costs_breakdown → from monthly tab transaction rows (new)
# tax, from, breakdown → preserved from existing JSON (manually annotated)
atl = []
for entry in monthly:
    m    = entry['m']
    prev = existing_atl.get(m, {})
    # Build sorted payees list for this month
    raw_payees = payees_by_month.get(m, {})
    payees_list = sorted(
        [{'name': k, 'amount': v} for k, v in raw_payees.items()],
        key=lambda x: -x['amount']
    )
    atl.append({
        'm':               m,
        'income':          income_by_month.get(m, 0.0),
        'costs':           atl_costs.get(m, 0.0),
        'costs_breakdown': costs_breakdown_by_month.get(m, []),
        'payees':          payees_list,
        'tax':             prev.get('tax', 0),
        'from':            prev.get('from', ''),
        'breakdown':       prev.get('breakdown', []),
    })

# ── Write JSON ────────────────────────────────────────────────────────────────
out = {
    'generated': datetime.now().isoformat(),
    'source':    os.path.basename(spreadsheet_path),
    'monthly':   monthly,
    'cats':      cats,
    'atl':       atl,
}
with open(JSON_OUT, 'w') as f:
    json.dump(out, f, indent=2)

# ── Summary ───────────────────────────────────────────────────────────────────
print(f"\nWritten: {JSON_OUT}")
print(f"  {len(monthly)} months · {monthly[0]['m']} – {monthly[-1]['m']}")

months_with_cost_breakdown = [m for m, bd in costs_breakdown_by_month.items() if bd]
if months_with_cost_breakdown:
    print(f"\nCost breakdowns found for: {', '.join(months_with_cost_breakdown)}")
else:
    print("\nNote: no cost breakdown column found in monthly tabs.")
    print("  Add a 'Project Costs' column to monthly tabs to enable per-payee tracking.")

if existing_atl:
    changes = []
    for entry in atl:
        old_income = existing_atl.get(entry['m'], {}).get('income')
        if old_income is not None and round(float(old_income), 2) != entry['income']:
            changes.append(f"  {entry['m']}: £{old_income:,.2f} → £{entry['income']:,.2f}")
    if changes:
        print("\nIncome changes vs previous JSON:")
        for c in changes: print(c)
    else:
        print("\nNo income changes vs previous JSON.")

print("\nNext step:")
print("  rm -f ~/push-films-website/.git/HEAD.lock && cd ~/push-films-website && "
      "git add dashboard_data.json && git commit -m 'Update dashboard data' && git push")
